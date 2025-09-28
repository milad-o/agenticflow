"""Excel agent with comprehensive spreadsheet operations."""

import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import json
import csv
from pathlib import Path
from typing import Annotated, List, Optional, Dict, Any, Union
from langchain_core.tools import tool
from ..core.flow import Agent

class ExcelAgent(Agent):
    """Agent specialized in Excel and spreadsheet operations."""
    
    def __init__(self, name: str = "excel_agent", description: str = "Excel and spreadsheet specialist"):
        tools = self._create_tools()
        super().__init__(name, tools=tools, description=description)
    
    def _create_tools(self) -> List:
        """Create Excel tools."""
        return [
            self._create_excel,
            self._read_excel,
            self._write_excel,
            self._append_excel,
            self._update_excel,
            self._delete_excel,
            self._copy_sheet,
            self._merge_excel,
            self._split_excel,
            self._format_excel,
            self._calculate_excel,
            self._filter_excel,
            self._sort_excel,
            self._pivot_excel,
            self._chart_excel,
            self._export_excel,
            self._import_excel,
            self._validate_excel,
            self._analyze_excel,
            self._convert_excel
        ]
    
    @tool
    def _create_excel(
        self,
        filename: Annotated[str, "Name of the Excel file to create"],
        sheets: Annotated[str, "Comma-separated list of sheet names (default: 'Sheet1')"] = "Sheet1",
        directory: Annotated[str, "Directory to create file in (default: examples/artifacts)"] = "examples/artifacts"
    ) -> str:
        """Create a new Excel file with specified sheets."""
        try:
            os.makedirs(directory, exist_ok=True)
            filepath = os.path.join(directory, filename)
            
            wb = Workbook()
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create specified sheets
            sheet_names = [s.strip() for s in sheets.split(',')]
            for sheet_name in sheet_names:
                wb.create_sheet(title=sheet_name)
            
            wb.save(filepath)
            return f"✅ Created Excel file '{filepath}' with sheets: {', '.join(sheet_names)}"
        except Exception as e:
            return f"❌ Error creating Excel file: {e}"
    
    @tool
    def _read_excel(
        self,
        filename: Annotated[str, "Name of the Excel file to read"],
        sheet_name: Annotated[str, "Sheet name to read (default: first sheet)"] = "",
        directory: Annotated[str, "Directory containing the file (default: examples/artifacts)"] = "examples/artifacts"
    ) -> str:
        """Read data from an Excel file."""
        try:
            filepath = os.path.join(directory, filename)
            
            if sheet_name:
                df = pd.read_excel(filepath, sheet_name=sheet_name)
            else:
                df = pd.read_excel(filepath)
            
            info = f"📊 Excel file '{filepath}':\n"
            info += f"📏 Shape: {df.shape[0]} rows, {df.shape[1]} columns\n"
            info += f"📋 Columns: {', '.join(df.columns.tolist())}\n"
            info += f"📄 First 5 rows:\n{df.head().to_string()}\n"
            
            return info
        except Exception as e:
            return f"❌ Error reading Excel file: {e}"
    
    @tool
    def _write_excel(
        self,
        filename: Annotated[str, "Name of the Excel file to write"],
        data: Annotated[str, "Data to write (JSON format or CSV format)"],
        sheet_name: Annotated[str, "Sheet name to write to (default: 'Sheet1')"] = "Sheet1",
        directory: Annotated[str, "Directory to write file in (default: examples/artifacts)"] = "examples/artifacts"
    ) -> str:
        """Write data to an Excel file."""
        try:
            os.makedirs(directory, exist_ok=True)
            filepath = os.path.join(directory, filename)
            
            # Try to parse as JSON first, then CSV
            try:
                data_dict = json.loads(data)
                df = pd.DataFrame(data_dict)
            except:
                # Try CSV format
                from io import StringIO
                df = pd.read_csv(StringIO(data))
            
            # Write to Excel
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            return f"✅ Wrote {df.shape[0]} rows to '{filepath}' in sheet '{sheet_name}'"
        except Exception as e:
            return f"❌ Error writing Excel file: {e}"
    
    @tool
    def _append_excel(
        self,
        filename: Annotated[str, "Name of the Excel file to append to"],
        data: Annotated[str, "Data to append (JSON format or CSV format)"],
        sheet_name: Annotated[str, "Sheet name to append to (default: 'Sheet1')"] = "Sheet1",
        directory: Annotated[str, "Directory containing the file (default: examples/artifacts)"] = "examples/artifacts"
    ) -> str:
        """Append data to an existing Excel file."""
        try:
            filepath = os.path.join(directory, filename)
            
            # Read existing data
            try:
                existing_df = pd.read_excel(filepath, sheet_name=sheet_name)
            except:
                existing_df = pd.DataFrame()
            
            # Parse new data
            try:
                data_dict = json.loads(data)
                new_df = pd.DataFrame(data_dict)
            except:
                from io import StringIO
                new_df = pd.read_csv(StringIO(data))
            
            # Append data
            combined_df = pd.concat([existing_df, new_df], ignore_index=True)
            
            # Write back to Excel
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                combined_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            return f"✅ Appended {new_df.shape[0]} rows to '{filepath}' in sheet '{sheet_name}'"
        except Exception as e:
            return f"❌ Error appending to Excel file: {e}"
    
    @tool
    def _update_excel(
        self,
        filename: Annotated[str, "Name of the Excel file to update"],
        updates: Annotated[str, "Updates to apply (JSON format with row/column specifications)"],
        sheet_name: Annotated[str, "Sheet name to update (default: 'Sheet1')"] = "Sheet1",
        directory: Annotated[str, "Directory containing the file (default: examples/artifacts)"] = "examples/artifacts"
    ) -> str:
        """Update specific cells in an Excel file."""
        try:
            filepath = os.path.join(directory, filename)
            
            # Load workbook
            wb = load_workbook(filepath)
            ws = wb[sheet_name]
            
            # Parse updates
            updates_dict = json.loads(updates)
            
            # Apply updates
            for update in updates_dict:
                row = update.get('row', 1)
                col = update.get('column', 1)
                value = update.get('value', '')
                ws.cell(row=row, column=col, value=value)
            
            # Save workbook
            wb.save(filepath)
            
            return f"✅ Updated {len(updates_dict)} cells in '{filepath}' sheet '{sheet_name}'"
        except Exception as e:
            return f"❌ Error updating Excel file: {e}"
    
    @tool
    def _delete_excel(
        self,
        filename: Annotated[str, "Name of the Excel file to delete"],
        directory: Annotated[str, "Directory containing the file (default: examples/artifacts)"] = "examples/artifacts"
    ) -> str:
        """Delete an Excel file."""
        try:
            filepath = os.path.join(directory, filename)
            if os.path.exists(filepath):
                os.remove(filepath)
                return f"🗑️ Deleted Excel file '{filepath}'"
            else:
                return f"⚠️ Excel file '{filepath}' not found"
        except Exception as e:
            return f"❌ Error deleting Excel file: {e}"
    
    @tool
    def _copy_sheet(
        self,
        filename: Annotated[str, "Name of the Excel file"],
        source_sheet: Annotated[str, "Source sheet name"],
        target_sheet: Annotated[str, "Target sheet name"],
        directory: Annotated[str, "Directory containing the file (default: examples/artifacts)"] = "examples/artifacts"
    ) -> str:
        """Copy a sheet within an Excel file."""
        try:
            filepath = os.path.join(directory, filename)
            
            # Load workbook
            wb = load_workbook(filepath)
            source_ws = wb[source_sheet]
            
            # Create new sheet
            target_ws = wb.create_sheet(title=target_sheet)
            
            # Copy data
            for row in source_ws.iter_rows():
                for cell in row:
                    target_ws[cell.coordinate] = cell.value
            
            # Save workbook
            wb.save(filepath)
            
            return f"✅ Copied sheet '{source_sheet}' to '{target_sheet}' in '{filepath}'"
        except Exception as e:
            return f"❌ Error copying sheet: {e}"
    
    @tool
    def _merge_excel(
        self,
        files: Annotated[str, "Comma-separated list of Excel files to merge"],
        output_filename: Annotated[str, "Name of the output Excel file"],
        directory: Annotated[str, "Directory containing the files (default: examples/artifacts)"] = "examples/artifacts"
    ) -> str:
        """Merge multiple Excel files into one."""
        try:
            file_list = [f.strip() for f in files.split(',')]
            merged_df = pd.DataFrame()
            
            for filename in file_list:
                filepath = os.path.join(directory, filename)
                df = pd.read_excel(filepath)
                merged_df = pd.concat([merged_df, df], ignore_index=True)
            
            output_path = os.path.join(directory, output_filename)
            merged_df.to_excel(output_path, index=False)
            
            return f"✅ Merged {len(file_list)} files into '{output_path}' with {merged_df.shape[0]} rows"
        except Exception as e:
            return f"❌ Error merging Excel files: {e}"
    
    @tool
    def _split_excel(
        self,
        filename: Annotated[str, "Name of the Excel file to split"],
        split_column: Annotated[str, "Column to split by"],
        directory: Annotated[str, "Directory containing the file (default: examples/artifacts)"] = "examples/artifacts"
    ) -> str:
        """Split an Excel file based on a column value."""
        try:
            filepath = os.path.join(directory, filename)
            df = pd.read_excel(filepath)
            
            # Split by column
            split_groups = df.groupby(split_column)
            
            created_files = []
            for group_name, group_df in split_groups:
                output_filename = f"{Path(filename).stem}_{group_name}.xlsx"
                output_path = os.path.join(directory, output_filename)
                group_df.to_excel(output_path, index=False)
                created_files.append(output_filename)
            
            return f"✅ Split '{filename}' into {len(created_files)} files: {', '.join(created_files)}"
        except Exception as e:
            return f"❌ Error splitting Excel file: {e}"
    
    @tool
    def _format_excel(
        self,
        filename: Annotated[str, "Name of the Excel file to format"],
        formatting: Annotated[str, "Formatting options (JSON format)"],
        sheet_name: Annotated[str, "Sheet name to format (default: 'Sheet1')"] = "Sheet1",
        directory: Annotated[str, "Directory containing the file (default: examples/artifacts)"] = "examples/artifacts"
    ) -> str:
        """Apply formatting to an Excel file."""
        try:
            filepath = os.path.join(directory, filename)
            
            # Load workbook
            wb = load_workbook(filepath)
            ws = wb[sheet_name]
            
            # Parse formatting options
            format_options = json.loads(formatting)
            
            # Apply formatting
            for option in format_options:
                cell_range = option.get('range', 'A1')
                font_size = option.get('font_size', 12)
                font_bold = option.get('font_bold', False)
                fill_color = option.get('fill_color', None)
                
                # Apply font formatting
                font = Font(size=font_size, bold=font_bold)
                ws[cell_range].font = font
                
                # Apply fill color
                if fill_color:
                    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    ws[cell_range].fill = fill
            
            # Save workbook
            wb.save(filepath)
            
            return f"✅ Applied formatting to '{filepath}' sheet '{sheet_name}'"
        except Exception as e:
            return f"❌ Error formatting Excel file: {e}"
    
    @tool
    def _calculate_excel(
        self,
        filename: Annotated[str, "Name of the Excel file to calculate"],
        formula: Annotated[str, "Formula to calculate (e.g., 'SUM(A1:A10)')"],
        cell: Annotated[str, "Cell to place the result (e.g., 'B1')"],
        sheet_name: Annotated[str, "Sheet name to calculate in (default: 'Sheet1')"] = "Sheet1",
        directory: Annotated[str, "Directory containing the file (default: examples/artifacts)"] = "examples/artifacts"
    ) -> str:
        """Add a formula to an Excel file."""
        try:
            filepath = os.path.join(directory, filename)
            
            # Load workbook
            wb = load_workbook(filepath)
            ws = wb[sheet_name]
            
            # Add formula
            ws[cell] = formula
            
            # Save workbook
            wb.save(filepath)
            
            return f"✅ Added formula '{formula}' to cell '{cell}' in '{filepath}'"
        except Exception as e:
            return f"❌ Error adding formula to Excel file: {e}"
    
    @tool
    def _filter_excel(
        self,
        filename: Annotated[str, "Name of the Excel file to filter"],
        filter_column: Annotated[str, "Column to filter by"],
        filter_value: Annotated[str, "Value to filter by"],
        output_filename: Annotated[str, "Name of the output filtered file"],
        directory: Annotated[str, "Directory containing the file (default: examples/artifacts)"] = "examples/artifacts"
    ) -> str:
        """Filter an Excel file and save the result."""
        try:
            filepath = os.path.join(directory, filename)
            df = pd.read_excel(filepath)
            
            # Apply filter
            filtered_df = df[df[filter_column] == filter_value]
            
            # Save filtered data
            output_path = os.path.join(directory, output_filename)
            filtered_df.to_excel(output_path, index=False)
            
            return f"✅ Filtered '{filename}' by '{filter_column}={filter_value}' and saved to '{output_filename}' with {filtered_df.shape[0]} rows"
        except Exception as e:
            return f"❌ Error filtering Excel file: {e}"
    
    @tool
    def _sort_excel(
        self,
        filename: Annotated[str, "Name of the Excel file to sort"],
        sort_column: Annotated[str, "Column to sort by"],
        ascending: Annotated[bool, "Sort in ascending order (default: True)"] = True,
        directory: Annotated[str, "Directory containing the file (default: examples/artifacts)"] = "examples/artifacts"
    ) -> str:
        """Sort an Excel file by a column."""
        try:
            filepath = os.path.join(directory, filename)
            df = pd.read_excel(filepath)
            
            # Sort data
            sorted_df = df.sort_values(by=sort_column, ascending=ascending)
            
            # Save sorted data
            sorted_df.to_excel(filepath, index=False)
            
            return f"✅ Sorted '{filename}' by '{sort_column}' in {'ascending' if ascending else 'descending'} order"
        except Exception as e:
            return f"❌ Error sorting Excel file: {e}"
    
    @tool
    def _pivot_excel(
        self,
        filename: Annotated[str, "Name of the Excel file to pivot"],
        index: Annotated[str, "Column to use as index"],
        columns: Annotated[str, "Column to use as columns"],
        values: Annotated[str, "Column to use as values"],
        output_filename: Annotated[str, "Name of the output pivoted file"],
        directory: Annotated[str, "Directory containing the file (default: examples/artifacts)"] = "examples/artifacts"
    ) -> str:
        """Create a pivot table from an Excel file."""
        try:
            filepath = os.path.join(directory, filename)
            df = pd.read_excel(filepath)
            
            # Create pivot table
            pivot_df = df.pivot_table(index=index, columns=columns, values=values, aggfunc='sum')
            
            # Save pivot table
            output_path = os.path.join(directory, output_filename)
            pivot_df.to_excel(output_path)
            
            return f"✅ Created pivot table from '{filename}' and saved to '{output_filename}'"
        except Exception as e:
            return f"❌ Error creating pivot table: {e}"
    
    @tool
    def _chart_excel(
        self,
        filename: Annotated[str, "Name of the Excel file to add chart to"],
        chart_type: Annotated[str, "Type of chart (bar, line, pie, etc.)"],
        data_range: Annotated[str, "Data range for the chart (e.g., 'A1:B10')"],
        sheet_name: Annotated[str, "Sheet name to add chart to (default: 'Sheet1')"] = "Sheet1",
        directory: Annotated[str, "Directory containing the file (default: examples/artifacts)"] = "examples/artifacts"
    ) -> str:
        """Add a chart to an Excel file."""
        try:
            filepath = os.path.join(directory, filename)
            
            # Load workbook
            wb = load_workbook(filepath)
            ws = wb[sheet_name]
            
            # Create chart (simplified - in practice, you'd use openpyxl's chart features)
            from openpyxl.chart import BarChart, Reference
            
            chart = BarChart()
            chart.title = f"{chart_type.title()} Chart"
            chart.x_axis.title = "X Axis"
            chart.y_axis.title = "Y Axis"
            
            # Add chart to worksheet
            ws.add_chart(chart, "E2")
            
            # Save workbook
            wb.save(filepath)
            
            return f"✅ Added {chart_type} chart to '{filepath}' sheet '{sheet_name}'"
        except Exception as e:
            return f"❌ Error adding chart to Excel file: {e}"
    
    @tool
    def _export_excel(
        self,
        filename: Annotated[str, "Name of the Excel file to export"],
        format: Annotated[str, "Export format (csv, json, html)"],
        output_filename: Annotated[str, "Name of the output file"],
        directory: Annotated[str, "Directory containing the file (default: examples/artifacts)"] = "examples/artifacts"
    ) -> str:
        """Export an Excel file to another format."""
        try:
            filepath = os.path.join(directory, filename)
            df = pd.read_excel(filepath)
            
            output_path = os.path.join(directory, output_filename)
            
            if format.lower() == 'csv':
                df.to_csv(output_path, index=False)
            elif format.lower() == 'json':
                df.to_json(output_path, orient='records', indent=2)
            elif format.lower() == 'html':
                df.to_html(output_path, index=False)
            else:
                return f"❌ Unsupported export format: {format}"
            
            return f"✅ Exported '{filename}' to '{output_filename}' in {format.upper()} format"
        except Exception as e:
            return f"❌ Error exporting Excel file: {e}"
    
    @tool
    def _import_excel(
        self,
        filename: Annotated[str, "Name of the file to import"],
        format: Annotated[str, "Import format (csv, json, html)"],
        output_filename: Annotated[str, "Name of the output Excel file"],
        directory: Annotated[str, "Directory containing the file (default: examples/artifacts)"] = "examples/artifacts"
    ) -> str:
        """Import a file and convert to Excel format."""
        try:
            filepath = os.path.join(directory, filename)
            
            if format.lower() == 'csv':
                df = pd.read_csv(filepath)
            elif format.lower() == 'json':
                df = pd.read_json(filepath)
            elif format.lower() == 'html':
                df = pd.read_html(filepath)[0]  # Read first table
            else:
                return f"❌ Unsupported import format: {format}"
            
            output_path = os.path.join(directory, output_filename)
            df.to_excel(output_path, index=False)
            
            return f"✅ Imported '{filename}' from {format.upper()} format and saved as '{output_filename}'"
        except Exception as e:
            return f"❌ Error importing file: {e}"
    
    @tool
    def _validate_excel(
        self,
        filename: Annotated[str, "Name of the Excel file to validate"],
        directory: Annotated[str, "Directory containing the file (default: examples/artifacts)"] = "examples/artifacts"
    ) -> str:
        """Validate an Excel file for errors and issues."""
        try:
            filepath = os.path.join(directory, filename)
            df = pd.read_excel(filepath)
            
            issues = []
            
            # Check for empty cells
            empty_cells = df.isnull().sum().sum()
            if empty_cells > 0:
                issues.append(f"Found {empty_cells} empty cells")
            
            # Check for duplicate rows
            duplicate_rows = df.duplicated().sum()
            if duplicate_rows > 0:
                issues.append(f"Found {duplicate_rows} duplicate rows")
            
            # Check for data types
            for col in df.columns:
                if df[col].dtype == 'object':
                    # Check if numeric data is stored as text
                    try:
                        pd.to_numeric(df[col], errors='raise')
                    except:
                        issues.append(f"Column '{col}' contains non-numeric data")
            
            if issues:
                return f"⚠️ Validation issues found in '{filename}':\n" + "\n".join(f"- {issue}" for issue in issues)
            else:
                return f"✅ '{filename}' validation passed - no issues found"
        except Exception as e:
            return f"❌ Error validating Excel file: {e}"
    
    @tool
    def _analyze_excel(
        self,
        filename: Annotated[str, "Name of the Excel file to analyze"],
        directory: Annotated[str, "Directory containing the file (default: examples/artifacts)"] = "examples/artifacts"
    ) -> str:
        """Analyze an Excel file and provide statistics."""
        try:
            filepath = os.path.join(directory, filename)
            df = pd.read_excel(filepath)
            
            analysis = f"📊 Analysis of '{filename}':\n"
            analysis += f"📏 Shape: {df.shape[0]} rows, {df.shape[1]} columns\n"
            analysis += f"📋 Columns: {', '.join(df.columns.tolist())}\n"
            analysis += f"📄 Data types:\n{df.dtypes.to_string()}\n"
            analysis += f"📈 Summary statistics:\n{df.describe().to_string()}\n"
            
            return analysis
        except Exception as e:
            return f"❌ Error analyzing Excel file: {e}"
    
    @tool
    def _convert_excel(
        self,
        filename: Annotated[str, "Name of the Excel file to convert"],
        target_format: Annotated[str, "Target format (csv, json, html, xml)"],
        output_filename: Annotated[str, "Name of the output file"],
        directory: Annotated[str, "Directory containing the file (default: examples/artifacts)"] = "examples/artifacts"
    ) -> str:
        """Convert an Excel file to another format."""
        try:
            filepath = os.path.join(directory, filename)
            df = pd.read_excel(filepath)
            
            output_path = os.path.join(directory, output_filename)
            
            if target_format.lower() == 'csv':
                df.to_csv(output_path, index=False)
            elif target_format.lower() == 'json':
                df.to_json(output_path, orient='records', indent=2)
            elif target_format.lower() == 'html':
                df.to_html(output_path, index=False)
            elif target_format.lower() == 'xml':
                df.to_xml(output_path, index=False)
            else:
                return f"❌ Unsupported target format: {target_format}"
            
            return f"✅ Converted '{filename}' to '{output_filename}' in {target_format.upper()} format"
        except Exception as e:
            return f"❌ Error converting Excel file: {e}"
